"""Chargeur unique des données AGM 2025 dans le cube.

    STORE_DSN=... python3 services/store/load_2025.py [--dry-run]

CE QUE CE SCRIPT EST, ET CE QU'IL N'EST PAS. C'est un chargement ponctuel,
autorisé explicitement pour FIDES, et non le début d'une chaîne d'ingestion.
Rien ici ne doit être réutilisé pour H2 : la cartographie source → modèle (#47)
et son taux de couverture calculé sont un tout autre sujet. Le marquer plutôt
que le sous-entendre évite qu'il devienne l'ingestion par accident.

CE QU'IL LIT. Le paquet physique AGM, qui est **confidentiel au titre de la
clause 9** et n'entre jamais dans le dépôt. Le script le lit depuis le disque
local et écrit dans la base ; il ne recopie rien dans un fichier suivi.

CE QU'IL ÉCRIT. Une cellule par (département, mois) de la feuille 3, une par
(produit, mois) de la feuille 7, et la part amont de chaque litre — le terme
qu'un modèle naïf perd entièrement, 22,8 % de la combustion.

L'IDEMPOTENCE. Les identifiants de cellule sont déterministes, dérivés de la
source. Relancer le chargement remplace au lieu d'empiler : un chargement joué
deux fois par mégarde doublerait l'inventaire, ce qui ne se verrait pas.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import psycopg
from psycopg.types.range import Range

import db

ROOT = Path(__file__).resolve().parents[2]
PACK = ROOT / "poc-data" / "AGM_PoC_Physical_Data_Pack_Completed.xlsx"
ASSIGNMENT = ROOT / "poc-data" / "agm-h1-subpost-assignment.json"
FIXTURE = ROOT / "site" / "static" / "engine" / "erp-fixture.json"

#: Feuille 9 du paquet. Cinq des six sont marqués « provisional » par AGM ;
#: seul le facteur de combustion du gazole est accepté, et seulement parce que
#: janvier réconcilie à −0,16 % contre leur propre classeur.
# LA CONVERSION A LIEU ICI, ET NULLE PART AILLEURS.
#
# La feuille 9 donne les facteurs par litre, parce que les bons de sortie
# comptent en litres. Le cube ne connaît que le mètre cube — et il ne stocke pas
# « kgCO2e/m3 » à côté du nombre, parce que l'unité d'activité est déjà celle de
# la cellule. Un facteur y est un NOMBRE : des kgCO2e par unité, et l'unité est
# dite une seule fois.
#
# C'est donc ici, à la frontière, que les unités d'origine existent encore et
# que la conversion se décide. Au-delà, il n'y a plus rien à convertir ni rien à
# vérifier.
SOURCE_FACTORS = {
    # facteur feuille 9, unité d'origine -> (facteur SI, unité SI)
    "diesel-combustion": (2.68, "L"),
    "diesel-upstream": (0.61, "L"),
    "explosive": (0.17, "kg"),
}

#: Ce que l'on sait ramener au SI. Une unité absente est REFUSÉE : convertir au
#: jugé produirait un nombre plausible et faux, que la signature figerait.
TO_SI = {
    "L": (1e3, "m3"),      # 1 kgCO2e/L = 1000 kgCO2e/m3
    "kg": (1.0, "kg"),
    "kWh": (1.0, "kWh"),
}


#: L'unité SI d'activité vers la DIMENSION qu'elle mesure.
#:
#: C'est la dimension qui est stockée, pas l'unité : sous l'hypothèse SI, elle
#: détermine l'unité sans laisser de choix — « volume » se lit m3, et divisé par
#: la durée de sa période, m3/s. Nommer l'unité par-dessus n'ajouterait qu'une
#: occasion de la contredire.
DIMENSION_OF = {"m3": "volume", "kg": "mass", "kWh": "energy"}


def si_factor(name: str) -> tuple[float, str]:
    """Le facteur ramené au SI, et l'unité d'activité qui va avec."""
    value, unit = SOURCE_FACTORS[name]
    if unit not in TO_SI:
        raise ValueError(f"unité de facteur non convertible : {unit}")
    scale, si_unit = TO_SI[unit]
    return value * scale, si_unit


def cell_metrology(name: str) -> tuple[str, str, float]:
    """La dimension, l'unité d'affichage, et le facteur qui va du SI vers elle.

    L'unité d'affichage est celle de la SOURCE — le litre des bons de sortie
    d'AGM — et elle ne sert qu'à relire la donnée brute sans compter les zéros.
    Aucun calcul ne la lit : la seule unité qui compte est celle du SI, et elle
    se déduit de la dimension.

    LE FACTEUR EST STOCKÉ PLUTÔT QUE DÉDUIT DU SYMBOLE, et c'est ce qui évite
    d'écrire un système d'unités. Savoir qu'une donnée « est en litres » ne sert
    à rien sans une table des symboles, de leurs préfixes et de leurs multiples,
    tenue juste. `display = SI x facteur` dit la même chose en un double.

    C'est le MÊME nombre que celui qui convertit le facteur d'émission, et pour
    la même raison : mille litres dans un mètre cube. `TO_SI` le porte une fois,
    et les deux usages le lisent là.
    """
    _, source_unit = SOURCE_FACTORS[name]
    scale, si_unit = TO_SI[source_unit]
    return DIMENSION_OF[si_unit], source_unit, scale


DIESEL_COMBUSTION, _ = si_factor("diesel-combustion")
DIESEL_UPSTREAM, _ = si_factor("diesel-upstream")
EXPLOSIVE, _ = si_factor("explosive")

DIESEL_DIMENSION, DIESEL_DISPLAY, DIESEL_SCALE = cell_metrology("diesel-combustion")
EXPLOSIVE_DIMENSION, EXPLOSIVE_DISPLAY, EXPLOSIVE_SCALE = cell_metrology("explosive")

#: Identifiants de la taxonomie servie, agm-h1-v2.
PART_COMBUSTION, PART_AMONT = 1, 2
CARAC_OPERATED, CARAC_PROCEDEED = 1, 2
SUBPOST_EXPLOSIVES = 1005

#: L'organisation de tête du client, racine de son sous-arbre.
#:
#: C'est elle qui rend le cube multi-client lisible : un client est le sous-arbre
#: suspendu à sa tête, et « compter par client » se dit « remonter les parents
#: jusqu'à la racine ». Les départements d'AGM portent les identifiants 1 à 34,
#: qui viennent du fixture ; la tête prend 100 pour qu'ajouter un département
#: n'entre jamais en collision avec elle.
#:
#: L'identité légale est ici et non dans le dépôt côté front : c'est une donnée
#: du client, elle vit dans SA taxonomie — celle-là même que le chiffrement des
#: dimensions couvrira le jour où D1 de l'issue #6 sera tranchée. En clair
#: aujourd'hui, comme les noms de départements, et pour la même raison.
HEAD = {
    "id": 100,
    "key": "AGM Inc.",
    "industrial": False,
    "legal_name": "AGM Inc",
    "jurisdiction": "Co-operative Republic of Guyana",
    "registered_office": ("3rd Floor R & S Mall Apartment District Track "
                          "JW Mandela Avenue, Durban Backlands, Georgetown, Guyana"),
    # Le domaine est celui de la mine, et elle le contrôle — décision 4 du
    # script de tournage, tranchée le 1er août. C'est ce DID que le front
    # inscrit comme émetteur des attestations d'origine.
    "did": "did:web:guygold.com",
}

#: Le Guyana est à UTC−4 toute l'année, sans heure d'été.
GUYANA = timezone(timedelta(hours=-4))


def month_range(label: str) -> Range:
    """« 2025-01 » -> le mois LOCAL, en bornes semi-ouvertes.

    Minuit à Georgetown, pas minuit à Greenwich. Une mine rapporte en jours
    guyaniens : découper sur des minuits Zulu décalerait chaque frontière de
    quatre heures et rangerait une nuit de production dans le mois suivant. À
    l'échelle d'un mois l'erreur est petite ; à la frontière d'un lot elle met
    du gazole dans la mauvaise barre, ce que le modèle prétend justement éviter.

    Semi-ouvertes parce que sinon une cellule serait comptée deux fois sur deux
    fenêtres adjacentes.
    """
    year, month = (int(x) for x in label.split("-"))
    start = datetime(year, month, 1, tzinfo=GUYANA)
    end = datetime(year + (month == 12), (month % 12) + 1, 1, tzinfo=GUYANA)
    return Range(start.astimezone(timezone.utc), end.astimezone(timezone.utc), "[)")


def read_pack():
    import openpyxl

    wb = openpyxl.load_workbook(PACK, data_only=True)
    fuel, explosives = [], []
    for r in wb["3. Fuel by consumer"].iter_rows(min_row=6, values_only=True):
        if r[1] and isinstance(r[3], (int, float)):
            fuel.append({"month": r[0], "department": r[1], "m3": float(r[3]) / 1000.0})
    for r in wb["7. Explosives"].iter_rows(min_row=6, values_only=True):
        if r[1] and isinstance(r[2], (int, float)):
            explosives.append({"month": r[0], "product": r[1], "kg": float(r[2])})
    return fuel, explosives


def organisation() -> dict[str, dict]:
    """Les départements, avec leurs identifiants — lus, jamais redevinés.

    Le fixture du front est la source unique. Attribuer les identifiants ici
    aussi produirait deux numérotations qui coïncideraient jusqu'au jour où un
    département serait ajouté, et la divergence serait silencieuse.

    Chacun est rattaché à l'organisation de tête : c'est ce rattachement, et lui
    seul, qui permet de compter par client.
    """
    fx = json.loads(FIXTURE.read_text("utf-8"))
    return {d["key"]: {**d, "parent": HEAD["id"]} for d in fx["organisation"]}


def synthetic_months() -> dict[str, str]:
    """Les mois absents du paquet et le mois d'où chacun se reconstitue.

    DÉCLARÉS PAR LE GÉNÉRATEUR, PAS REDEVINÉS ICI. `make_fixture.py` a déjà
    tranché quels mois manquent — la fenêtre de production commence en février
    parce que janvier puiserait dans un décembre 2024 que le paquet ne contient
    pas — et il l'écrit dans `model.syntheticMonths`. Recalculer la même chose
    depuis le classeur donnerait deux vérités qui coïncideraient jusqu'au jour
    où l'une changerait.

    Le donneur est le même mois de l'année suivante : décembre 2024 se reconstitue
    depuis décembre 2025, ce qui est la seule saisonnalité que douze mois de
    données permettent d'invoquer.
    """
    fx = json.loads(FIXTURE.read_text("utf-8"))
    missing = fx.get("model", {}).get("syntheticMonths", [])
    return {m: f"{int(m.split('-')[0]) + 1}-{m.split('-')[1]}" for m in missing}


def reconstruct(rows, donors):
    """Ajoute les mois manquants, recopiés de leur donneur et marqués MISSING.

    LA CELLULE EXISTE, ET ELLE LE DIT. Ne rien charger pour décembre 2024
    laisserait un trou qu'aucun dénombrement ne verrait : une couverture calculée
    sur les cellules présentes vaudrait 100 % en ignorant le mois absent. Charger
    la reconstitution SANS la marquer serait pire — le trou deviendrait invisible
    tout en pesant sur les chiffres publiés.

    Le seul choix honnête est de porter la valeur et l'aveu ensemble.
    """
    made = []
    for missing, donor in donors.items():
        source = [r for r in rows if r["month"] == donor]
        if not source:
            print(f"  aucun donneur {donor} pour {missing} — mois laissé vide",
                  file=sys.stderr)
            continue
        made += [{**r, "month": missing, "coverage": "MISSING"} for r in source]
    return rows + made


def period_seconds(period: Range) -> float:
    """La durée d'un intervalle, en secondes.

    C'est le diviseur qui fait d'une quantité un débit. Un intervalle vide est
    impossible ici — la contrainte `period_not_empty` du schéma le refuse — donc
    il n'y a pas de division par zéro à craindre, et rien à contrôler deux fois.
    """
    return (period.upper - period.lower).total_seconds()


def build_cells(fuel, explosives, assignment, org):
    """La correspondance département → sous-poste vient du fichier d'affectation,
    relu et non redeviné : c'est lui qui porte les 14 départements marqués
    « needs AGM confirmation », et il a été vérifié séparément."""
    by_department = {e["department"]: e for e in assignment["emissionBearing"]["fuel"]}
    cells = []

    for row in fuel:
        spec = by_department.get(row["department"])
        if spec is None:
            print(f"  département inconnu, ignoré : {row['department']}", file=sys.stderr)
            continue
        sub_post = {"CombustiblesFossiles": 1000, "FretInterne": 1002}[spec["subPost"]]
        dept = org.get(row["department"])
        if dept is None:
            print(f"  département hors taxonomie, ignoré : {row['department']}", file=sys.stderr)
            continue
        slug = row["department"].replace(" ", "_")[:40]
        period = month_range(row["month"])
        # LA DIVISION QUI FAIT LE DÉBIT, et elle a lieu ici, à l'ingestion, au
        # même endroit que la conversion des unités et pour la même raison :
        # c'est le dernier point où la donnée source existe encore telle que la
        # mine l'a écrite. Au-delà, il n'y a qu'un débit sur un intervalle.
        flux = row["m3"] / period_seconds(period)
        # Deux cellules par litre : la combustion, et l'amont. Charger seulement
        # la première perdrait 22,8 % de l'empreinte sans que rien ne le signale.
        for part, factor, tag in ((PART_COMBUSTION, DIESEL_COMBUSTION, "comb"),
                                  (PART_AMONT, DIESEL_UPSTREAM, "amont")):
            cells.append({
                "id": f"d/{row['month']}/{slug}/{tag}",
                "period": period,
                "entity_id": dept["id"],
                "sub_post": sub_post, "part_type": part,
                "caracterisation": CARAC_OPERATED,
                "flux": flux,
                "dimension": DIESEL_DIMENSION, "display_unit": DIESEL_DISPLAY,
                "display_scale": DIESEL_SCALE,
                "factor": factor,
                "origin": "MEASURED",
                "coverage": row.get("coverage", "COMPLETE"),
            })

    blast_dept = org["Sinohydro"]["id"]
    for row in explosives:
        period = month_range(row["month"])
        cells.append({
            "id": f"x/{row['month']}/{row['product'].replace(' ', '_')}",
            "period": period,
            # Le paquet donne les explosifs par produit et par mois, jamais par
            # département : ils vont au département de minage. Les répartir
            # entre plusieurs inventerait une ventilation que personne n'a.
            "entity_id": blast_dept,
            "sub_post": SUBPOST_EXPLOSIVES, "part_type": None,
            "caracterisation": CARAC_PROCEDEED,
            "flux": row["kg"] / period_seconds(period),
            "dimension": EXPLOSIVE_DIMENSION, "display_unit": EXPLOSIVE_DISPLAY,
            "display_scale": EXPLOSIVE_SCALE,
            "factor": EXPLOSIVE,
            "origin": "MEASURED",
            "coverage": row.get("coverage", "COMPLETE"),
        })

    return cells


def load(conn, cells, org) -> None:
    db.apply_schema(conn)
    # La taxonomie d'organisation. Les noms sont en clair PROVISOIREMENT : ce
    # sont eux que le chiffrement des dimensions couvrira. Le client n'en connaît
    # déjà que les entiers.
    #
    # LA TÊTE AVANT LES DÉPARTEMENTS, et pas par habitude : `parent` référence
    # `entity(id)`, donc insérer un département avant sa tête viole la clé
    # étrangère et fait avorter le chargement entier.
    conn.execute(
        """INSERT INTO entity (id, label, industrial, legal_name, jurisdiction,
                               registered_office, did)
                VALUES (%(id)s, %(key)s, %(industrial)s, %(legal_name)s,
                        %(jurisdiction)s, %(registered_office)s, %(did)s)
           ON CONFLICT (id) DO UPDATE SET label = EXCLUDED.label,
                industrial = EXCLUDED.industrial, legal_name = EXCLUDED.legal_name,
                jurisdiction = EXCLUDED.jurisdiction,
                registered_office = EXCLUDED.registered_office,
                did = EXCLUDED.did""",
        HEAD,
    )
    with conn.cursor() as cur:
        cur.executemany(
            """INSERT INTO entity (id, label, parent, industrial)
                    VALUES (%(id)s, %(key)s, %(parent)s, %(industrial)s)
               ON CONFLICT (id) DO UPDATE SET label = EXCLUDED.label,
                                              parent = EXCLUDED.parent,
                                              industrial = EXCLUDED.industrial""",
            list(org.values()),
        )

    with conn.cursor() as cur:
        cur.executemany(
            """INSERT INTO cell (id, period, entity_id, sub_post, part_type, caracterisation,
                                 flux, dimension, display_unit, display_scale,
                                 factor, origin, coverage)
                    VALUES (%(id)s, %(period)s, %(entity_id)s, %(sub_post)s, %(part_type)s,
                            %(caracterisation)s, %(flux)s, %(dimension)s, %(display_unit)s,
                            %(display_scale)s, %(factor)s, %(origin)s, %(coverage)s)
               ON CONFLICT (id) DO UPDATE SET
                    period = EXCLUDED.period, entity_id = EXCLUDED.entity_id,
                    sub_post = EXCLUDED.sub_post, part_type = EXCLUDED.part_type,
                    caracterisation = EXCLUDED.caracterisation,
                    flux = EXCLUDED.flux, dimension = EXCLUDED.dimension,
                    display_unit = EXCLUDED.display_unit,
                    display_scale = EXCLUDED.display_scale,
                    factor = EXCLUDED.factor, origin = EXCLUDED.origin,
                    coverage = EXCLUDED.coverage""",
            cells,
        )


def sql_literal(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return repr(value)
    # Les apostrophes des noms de département — « BH Contractor - Ping An » n'en
    # a pas, mais rien ne garantit que le prochain n'en aura pas.
    return "'" + str(value).replace("'", "''") + "'"


def emit_sql(cells, org) -> None:
    """Le chargement, en SQL, sur stdout.

    Une seule transaction : un chargement à moitié appliqué laisserait un cube
    dont personne ne saurait dire s'il est complet.
    """
    print("BEGIN;")
    # La tête d'abord : `parent` la référence.
    print("INSERT INTO entity (id, label, industrial, legal_name, jurisdiction, "
          "registered_office, did) VALUES "
          f"({HEAD['id']}, {sql_literal(HEAD['key'])}, {sql_literal(HEAD['industrial'])}, "
          f"{sql_literal(HEAD['legal_name'])}, {sql_literal(HEAD['jurisdiction'])}, "
          f"{sql_literal(HEAD['registered_office'])}, {sql_literal(HEAD['did'])}) "
          "ON CONFLICT (id) DO UPDATE SET label = EXCLUDED.label, "
          "industrial = EXCLUDED.industrial, legal_name = EXCLUDED.legal_name, "
          "jurisdiction = EXCLUDED.jurisdiction, "
          "registered_office = EXCLUDED.registered_office, did = EXCLUDED.did;")
    for d in org.values():
        print(f"INSERT INTO entity (id, label, parent, industrial) VALUES "
              f"({d['id']}, {sql_literal(d['key'])}, {sql_literal(d['parent'])}, "
              f"{sql_literal(d['industrial'])}) "
              "ON CONFLICT (id) DO UPDATE SET label = EXCLUDED.label, "
              "parent = EXCLUDED.parent, industrial = EXCLUDED.industrial;")
    for c in cells:
        lo = c["period"].lower.isoformat()
        hi = c["period"].upper.isoformat()
        print(
            "INSERT INTO cell (id, period, entity_id, sub_post, part_type, caracterisation,"
            " flux, dimension, display_unit, display_scale, factor, origin, coverage)"
            " VALUES ("
            f"{sql_literal(c['id'])}, "
            f"tstzrange({sql_literal(lo)}::timestamptz, {sql_literal(hi)}::timestamptz, '[)'), "
            f"{c['entity_id']}, {sql_literal(c['sub_post'])}, {sql_literal(c['part_type'])}, "
            f"{c['caracterisation']}, {c['flux']!r}, {sql_literal(c['dimension'])}, "
            f"{sql_literal(c['display_unit'])}, {c['display_scale']!r}, {c['factor']!r}, "
            f"{sql_literal(c['origin'])}, {sql_literal(c['coverage'])}) "
            # Toutes les colonnes, sans exception : un rechargement qui change
            # de dimension doit changer la dimension. En omettre une laisse une
            # valeur neuve sous une étiquette ancienne.
            "ON CONFLICT (id) DO UPDATE SET period = EXCLUDED.period, "
            "entity_id = EXCLUDED.entity_id, sub_post = EXCLUDED.sub_post, "
            "part_type = EXCLUDED.part_type, caracterisation = EXCLUDED.caracterisation, "
            "flux = EXCLUDED.flux, dimension = EXCLUDED.dimension, "
            "display_unit = EXCLUDED.display_unit, display_scale = EXCLUDED.display_scale, "
            "factor = EXCLUDED.factor, origin = EXCLUDED.origin, "
            "coverage = EXCLUDED.coverage;"
        )
    print("COMMIT;")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                    help="compte et résume sans écrire — le défaut serait dangereux dans l'autre sens")
    ap.add_argument("--sql", action="store_true",
                    help="émet le SQL sur stdout au lieu de se connecter ; le paquet AGM reste ici")
    args = ap.parse_args()

    if not PACK.exists():
        print(f"paquet AGM introuvable : {PACK}", file=sys.stderr)
        print("il est confidentiel (clause 9) et n'est pas dans le dépôt.", file=sys.stderr)
        return 1

    fuel, explosives = read_pack()
    assignment = json.loads(ASSIGNMENT.read_text("utf-8"))
    org = organisation()

    # Les mois absents entrent AVANT la construction : ils produisent de vraies
    # cellules, avec un vrai identifiant déterministe, qui disent seulement
    # d'où elles viennent.
    donors = synthetic_months()
    fuel = reconstruct(fuel, donors)
    explosives = reconstruct(explosives, donors)

    cells = build_cells(fuel, explosives, assignment, org)
    missing = sum(1 for c in cells if c["coverage"] == "MISSING")

    # Le résumé reparle en quantités, parce qu'un débit de gazole en m3/s ne se
    # relit pas. C'est exactement l'opération que le moteur fait ensuite : débit
    # multiplié par la durée de sa période.
    quantity = lambda c: c["flux"] * period_seconds(c["period"])
    m3 = sum(quantity(c) for c in cells
             if c["dimension"] == "volume" and c["part_type"] == PART_COMBUSTION)
    tonnes = sum(quantity(c) * c["factor"] for c in cells) / 1000

    # Le résumé est un diagnostic, pas une donnée : il va sur stderr, sinon il
    # se mêle au SQL quand celui-ci part dans un tuyau.
    industrial = sum(1 for d in org.values() if d["industrial"])
    say = lambda m: print(m, file=sys.stderr)
    say(f"{len(cells)} cellules — {m3:,.0f} m3 de gazole, {tonnes:,.0f} tCO2e au total")
    say(f"  organisation : {HEAD['key']} + {len(org)} départements, dont {industrial} industriels")
    say(f"  couverture : {missing} MISSING ({', '.join(donors) or 'aucun mois reconstitué'})")
    say(f"  affectation : {assignment['version']} ({assignment['status'].split(' - ')[0]})")

    if args.dry_run:
        say("\n--dry-run : rien n'a été écrit")
        return 0

    if args.sql:
        # Le classeur AGM est confidentiel au titre de la clause 9 : il ne quitte
        # pas ce poste. Seules les données dérivées traversent, par stdin, et
        # rien ne s'écrit sur le système de fichiers de la cible — c'est la
        # doctrine de deploy/, elle vaut ici aussi.
        emit_sql(cells, org)
        return 0

    with db.connect() as conn:
        load(conn, cells, org)
        n = conn.execute("SELECT count(*) AS n FROM cell").fetchone()["n"]
    say(f"\nchargé. {n} cellules dans le cube.")
    return 0


if __name__ == "__main__":
    sys.exit(main())

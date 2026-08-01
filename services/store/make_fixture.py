"""Fabrique les données ERP du front : taxonomie d'organisation, procédé, lots, barres.

    python3 services/store/make_fixture.py

LE FUSEAU N'EST PAS UN DÉTAIL. Une mine du Guyana rapporte en jours guyaniens.
Découper les mois sur des minuits Zulu décale chaque frontière de quatre heures
et attribue au mois suivant une nuit de production. À l'échelle d'un lot mensuel
l'erreur est petite ; à la frontière d'un lot elle range du gazole dans la
mauvaise barre, et c'est précisément ce que le modèle prétend faire correctement.
Le Guyana est à UTC−4 toute l'année, sans heure d'été : un début de mois local
est donc 04:00Z le premier du mois.

LES DÉPARTEMENTS SONT DES ENTIERS. Ils sont déclarés dans une taxonomie
d'organisation, statique en base pour le PoC. Le client n'en connaît que les
index : un identifiant entier ne divulgue pas un organigramme, et le jour où la
table des noms sera chiffrée, rien du front ne changera.

LE PROCÉDÉ EST CODÉ UNE FOIS. Tous les lots suivent le même chemin ; le répéter
douze fois serait douze occasions de diverger. Un lot référence un procédé et
porte son propre paquet de données.

LE GAZOLE VIENT DES SORTIES, PAS DES FACTURES. La feuille 3 est « Diesel Issue »
depuis « Consumption 2025 » : du carburant sorti vers un engin, donc brûlé. Le
gazole facturé (jeu D-05, mouvement 109) ne porte que de l'amont tant qu'il dort
dans une cuve, et cet amont-là est hors périmètre. D-05 ne sert qu'au bilan
matière.

CE QUI EST SIMULÉ. Le registre de coulée d'AGM — G-01 — est encore partiel : la
date de coulée, l'identifiant de barre, le poids et le titre sont fabriqués. Les
onces par mois, les départements, le gazole et les explosifs viennent du paquet
réel. Décembre 2024, qu'aucun paquet ne couvre, est une copie de décembre 2025 :
les barres coulées début janvier doivent puiser quelque part, et cette copie ne
sert nulle part ailleurs.
"""

from __future__ import annotations

import json
import secrets
from datetime import datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
PACK = ROOT / "poc-data" / "AGM_PoC_Physical_Data_Pack_Completed.xlsx"
ASSIGNMENT = ROOT / "poc-data" / "agm-h1-subpost-assignment.json"
OUT = ROOT / "site" / "static" / "engine" / "erp-fixture.json"

TOTAL_BARS = 378

#: Le Guyana ne pratique pas l'heure d'été : un décalage fixe suffit, et évite
#: d'embarquer une base de fuseaux dans le navigateur.
GUYANA = timezone(timedelta(hours=-4))

#: Départements que la matière traverse. Les autres — camp, HSSE, informatique,
#: entrepôt, géologie — soutiennent la production sans être sur le chemin d'un
#: lot : leurs émissions deviennent le non-alloué.
INDUSTRIAL = {
    "Power Generation", "UG-Power Gen", "Sinohydro", "Mobile Maintenance",
    "UG-Mining (GMYM)", "UG-Mining (HDLS)", "UG-Mining Project", "Mine Operation",
    "Mill General Ops", "Mill Maintenance", "Buckhall", "BH Road Maint",
    "BH Contractor - Ping An", "BH Contractor - DTL", "BH Contractor - Yano",
    "BH Contractor - Morrison", "BH Contractor - SMS", "Tapir",
    "Fuel transport -DTL", "Fuel transport -SIR",
}

#: Le procédé, en étapes. Chaque étape groupe les départements qui l'exécutent,
#: et c'est à une étape que se rattache une consommation d'explosifs.
STEPS = [
    ("drill-blast", ["Sinohydro", "Mine Operation", "BH Road Maint"], True),
    ("haul", ["BH Contractor - Ping An", "BH Contractor - DTL", "BH Contractor - Yano",
              "BH Contractor - Morrison", "BH Contractor - SMS", "Buckhall",
              "Fuel transport -DTL", "Fuel transport -SIR"], False),
    ("underground", ["UG-Mining (GMYM)", "UG-Mining (HDLS)", "UG-Mining Project",
                     "UG-Power Gen"], False),
    ("mill", ["Mill General Ops", "Mill Maintenance"], False),
    ("power", ["Power Generation", "Tapir"], False),
    ("maintenance", ["Mobile Maintenance"], False),
]


def month_bounds(label: str) -> tuple[str, str]:
    """« 2025-03 » -> ses deux instants LOCAUX, exprimés en UTC.

    Minuit à Georgetown, pas minuit à Greenwich.
    """
    year, month = (int(x) for x in label.split("-"))
    start = datetime(year, month, 1, tzinfo=GUYANA)
    end = datetime(year + (month == 12), (month % 12) + 1, 1, tzinfo=GUYANA)
    iso = lambda d: d.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
    return iso(start), iso(end)


def previous_month(label: str) -> str:
    year, month = (int(x) for x in label.split("-"))
    return f"{year - (month == 1)}-{12 if month == 1 else month - 1:02d}"


def bars_per_month(ounces: dict[str, float], window: list[str]) -> dict[str, int]:
    """Au prorata des onces, par plus grand reste.

    Arrondir chaque mois indépendamment donnerait 374 ou 381 barres, et un total
    qui ne tombe pas juste se remarque tout de suite dans une démonstration.
    """
    total = sum(ounces[m] for m in window)
    raw = {m: TOTAL_BARS * ounces[m] / total for m in window}
    counts = {m: int(v) for m, v in raw.items()}
    for m, _ in sorted(raw.items(), key=lambda kv: kv[1] - int(kv[1]), reverse=True)[
        : TOTAL_BARS - sum(counts.values())
    ]:
        counts[m] += 1
    return counts


def main() -> int:
    import openpyxl

    wb = openpyxl.load_workbook(PACK, data_only=True)

    ounces = {}
    for r in wb["6. Production"].iter_rows(min_row=6, values_only=True):
        if r[0] and isinstance(r[7], (int, float)):
            ounces[r[0]] = float(r[7])

    fuel: dict[str, dict[str, float]] = {}
    for r in wb["3. Fuel by consumer"].iter_rows(min_row=6, values_only=True):
        if r[1] and isinstance(r[3], (int, float)):
            fuel.setdefault(r[0], {})[r[1]] = float(r[3])

    explosives: dict[str, dict[str, float]] = {}
    for r in wb["7. Explosives"].iter_rows(min_row=6, values_only=True):
        if r[1] and isinstance(r[2], (int, float)):
            explosives.setdefault(r[0], {})[r[1]] = float(r[2])

    assignment = json.loads(ASSIGNMENT.read_text("utf-8"))
    names = [e["department"] for e in assignment["emissionBearing"]["fuel"]]

    # La taxonomie d'organisation. Les identifiants sont stables — l'ordre du
    # fichier d'affectation, lui-même par part décroissante — et le client ne
    # verra jamais que ces entiers.
    dept_id = {name: i + 1 for i, name in enumerate(names)}
    organisation = [
        {"id": dept_id[n], "key": n, "industrial": n in INDUSTRIAL} for n in names
    ]

    process = {
        "id": 1,
        "key": "aurora-dore-v1",
        "steps": [
            {"id": i + 1, "key": key,
             "departments": sorted(dept_id[n] for n in depts if n in dept_id),
             "consumesExplosives": blast}
            for i, (key, depts, blast) in enumerate(STEPS)
        ],
    }
    blast_step = next(s["id"] for s in process["steps"] if s["consumesExplosives"])

    source_of = {m: previous_month(m) for m in sorted(ounces)}
    synthetic = sorted({m for m in source_of.values() if m not in fuel})
    for missing in synthetic:
        donor = f"{int(missing.split('-')[0]) + 1}-12"
        fuel[missing] = dict(fuel[donor])
        explosives[missing] = dict(explosives.get(donor, {}))

    window = sorted(ounces)
    counts = bars_per_month(ounces, window)

    lots, bars = [], []
    for pour_month in window:
        source = source_of[pour_month]
        start, end = month_bounds(source)
        lot_id = f"LOT-{source}"

        lots.append({
            "id": lot_id,
            "process": process["id"],
            "productionMonth": source,
            "pourMonth": pour_month,
            "period": {"start": start, "end": end},
            "syntheticSource": source in synthetic,
            # Le paquet de données du lot. Un seul procédé en H1, donc un seul
            # paquet possible : gazole par département, explosifs par produit.
            "dataPackage": {
                "diesel": [{"department": dept_id[d], "litres": round(v, 3)}
                           for d, v in sorted(fuel.get(source, {}).items()) if d in dept_id],
                "explosives": [{"step": blast_step, "product": p, "kg": round(v, 3)}
                               for p, v in sorted(explosives.get(source, {}).items())],
            },
        })

        oz_each = ounces[pour_month] / counts[pour_month]
        pour_start, _ = month_bounds(pour_month)
        for n in range(1, counts[pour_month] + 1):
            bars.append({
                # Aléa de 128 bits : une numérotation ordonnée divulguerait le
                # rythme de production.
                "subjectId": f"urn:aurora:dore:{secrets.token_hex(16)}",
                "internalId": f"AUR-{pour_month.replace('-', '')}-{n:03d}",
                "lot": lot_id,
                "pouredAt": pour_start,
                "ounces": round(oz_each, 2),
                "weightKg": round(oz_each * 0.0311034768 / 0.92, 3),
                "assay": 0.92,
            })

    fixture = {
        "simulated": True,
        "eventModel": "simulated-v1",
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        # Des valeurs, pas des phrases : ce qui s'affiche se traduit dans les
        # gabarits, et un fichier de données n'a pas de langue.
        "model": {
            "timezone": "America/Guyana",
            "utcOffset": "-04:00",
            "cycleMonths": 1,
            "pourAt": "month-start",
            "sourcing": "previous-month",
            "unallocatedRule": "unallocated/bars-poured-same-month",
            "dieselSource": "issues-to-equipment",
            "syntheticMonths": synthetic,
        },
        "simulatedFields": ["pouredAt", "internalId", "weightKg", "assay"],
        "organisation": organisation,
        "process": process,
        "lots": lots,
        "bars": bars,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(fixture, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"{len(bars)} barres, {len(lots)} lots, {len(organisation)} départements")
    print(f"  fenêtre {window[0]}..{window[-1]}, mois synthétiques : {synthetic or 'aucun'}")
    print(f"  bornes locales : {lots[0]['period']['start']} .. {lots[0]['period']['end']}")
    print(f"  -> {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
